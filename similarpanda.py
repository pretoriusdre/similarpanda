# -*- coding: utf-8 -*-
"""
Created on Tue Dec 14 20:00:39 2021

@author: Andre Pretorius
"""


from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import numpy as np
from datetime import date
import os # Used in the interface file
from pathlib import Path

START_ROW = 8
START_COL = 2
CUSTOM_LINE = r'=HYPERLINK("https://github.com/pretoriusdre/similarpanda")'

class SimilarPanda:
    """A class to check for differences between pandas Dataframes, and generate a report.

    Inputs:
        new: DataFrame
        old: DataFrame
        key_column: A reference to a common DataFrame column. If omitted, the
            data will be matched on the row position.

    Returns:
        Use output_excel to generate an Excel report.
    """
    
    def __init__(self,
                 new,
                 old,
                 key_column=None,
                 df_new_title='Not specified',
                 df_old_title='Not specified'
                 ):
        
        self.df_new_title = df_new_title
        self.df_old_title = df_old_title
        
        self.df_changes, (self.added_cols, self. deleted_cols) = (
            self.compare_dataframe_changes(new, old, key_column)
            )

    def compare_dataframe_changes(self, new, old, key_column=None):
        """ A function to compare two dataframes and return a summary of the changes
        Inputs:
        new: new dataframe
        old: old dataframe
        key_column: a column which is used to to match the data rows. If omitted, the rows will be matched by position.
        
        Returns:
        A dataframe with two columns to describe the changed cells across each row.
        A list of added columns
        A list of deleted columns
        """
        # new, old are dataframes
        # key_column is a dataframe column
        
        new = new.copy()
        old = old.copy()

        def get_row_status(record):
            """ A helper function to create a summary of whether a row has changed.
            Applied to each row in the dataframe"""
            explain = {'left_only': 'Row added', 'right_only': 'Row deleted', 'both': '-'}
            status = explain[record['_merge']]
            if status == '-':
                if len(record['Value changes']) > 0:
                    status = 'Row updated'
            return status   

        def get_row_changes(record):
            """A helper function to create a dictionary describing the changes in each row.
            Applied to each row in the dataframe"""
            changes_dict = {}
            for col in cols_reduced:
                
                old_val = None
                if col in old_cols:
                    old_val = record[col + '_old']
                
                if pd.isna(old_val):
                    old_val = None
                if record[col] != old_val:
                    new_val = record[col]
                    if pd.isna(new_val):
                        new_val = None
                    if (old_val is not None) or (new_val is not None):
                        changes_dict[col] = {old_val: new_val}
            return changes_dict

        new_cols = list(new.columns)
        old_cols = list(old.columns)
        
        added_cols = [col for col in new_cols if col not in old_cols]
        deleted_cols = [col for col in old_cols if col not in new_cols]

        #  Exclude the key column, as this will always match
        cols_reduced = [col for col in new_cols if col != key_column]
    
        old['row_id'] = np.arange(old.shape[0])
        new['row_id'] = np.arange(new.shape[0])
        
        if key_column is None:
            key_column = 'row_id'
        
        df_differences = pd.merge(new, old, on=key_column, suffixes=('', '_old'), how='outer', indicator=True)
        df_differences.drop('row_id', axis=1)
        df_differences['Value changes'] = df_differences.apply(get_row_changes, axis=1)
        df_differences['Row status'] = df_differences.apply(get_row_status, axis=1)
        cols_to_return = new_cols
        cols_to_return.extend(deleted_cols)
        cols_to_return.extend(['Row status', 'Value changes'])
        
        return df_differences[cols_to_return], (added_cols, deleted_cols)

    def output_excel(self, output_file, worksheet_title='Data changes'):
        """Generates an output Excel file to explain the data changes
        Makes use of inbuilt styles: 'Good', 'Bad', and 'Neutral'. Might need to be changed in other languages"""
        
        def xlref(row, column, zero_indexed=True):
            if zero_indexed:
                row += 1
                column += 1
            return get_column_letter(column) + str(row)

        cols = self.df_changes.columns
        wb = Workbook()
        ws1 = wb.active
        ws1.title = worksheet_title

        for col_index, col in enumerate(cols):
            cell = ws1.cell(column=(col_index + START_COL), row=START_ROW)
            cell.value = col
            if col in self.added_cols:
                cell.style = 'Good'
            if col in self.deleted_cols:
                cell.style = 'Bad'
    
        for row_index, record in self.df_changes.iterrows():

            for col_index, col in enumerate(cols):
                val_to_print = record[col]
                if type(val_to_print) == dict:
                    val_to_print = str(val_to_print).replace('},', '},\n')
                cell = ws1.cell(column=(col_index + START_COL), row=(row_index + START_ROW + 1))
                cell.value = val_to_print
                if col in record['Value changes'].keys():
                    cell.style = 'Neutral'

                if  col =='Value changes' and len(record['Value changes']) > 0:
                    cell.style = 'Neutral'
                    
                if col in self.added_cols:
                    cell.style = 'Good'
                if col in self.deleted_cols:
                    cell.style = 'Bad'
                
                if record['Row status'] == 'Row added':
                    cell.style = 'Good'
                if record['Row status'] == 'Row deleted':
                    cell.style = 'Bad'
                if cell.value == 'Row updated':
                    cell.style = 'Neutral'

        ws1.cell(column=2, row=2).value = 'Data comparison tool'
        for col_ref in range(2, 6):
            ws1.cell(column=col_ref, row=2).style = 'Headline 1'
        ws1.cell(column=2, row=3).value = 'New:'
        ws1.cell(column=3, row=3).value = self.df_new_title

        ws1.cell(column=2, row=4).value = 'Old:'
        ws1.cell(column=3, row=4).value = self.df_old_title

        ws1.cell(column=2, row=5).value = 'Generated:'
        ws1.cell(column=3, row=5).value = date.today().strftime("%Y-%m-%d")

        ws1.cell(column=2, row=6).value = CUSTOM_LINE
        ws1.cell(column=2, row=6).style = 'Hyperlink'

        table_style = TableStyleInfo(
            name='TableStyleMedium9',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False)

        top_left_cell = xlref(START_ROW, START_COL, zero_indexed=False)
        top_right_cell = xlref(START_ROW + len(self.df_changes), START_COL + len(cols) - 1, zero_indexed=False)
        table_ref = top_left_cell + ':' + top_right_cell
        tab = Table(displayName='TableDifferences', ref=table_ref)
        tab.tableStyleInfo = table_style    
        ws1.add_table(tab)
        
        wb.save(filename=output_file)


def get_all_tables_in_excel(filename):
    """ A helper function to extract all Named DataTables from an Excel file,
    and store these in a dictionary (key = Table name, value = Dataframe)"""
    
    wb = load_workbook(filename, data_only=True)
    mapping = {}

    for ws in wb.worksheets:

        for entry, data_boundary in ws.tables.items():
            # Parse the data within the ref boundary
            data = ws[data_boundary]
            # The inner list comprehension gets the values for each cell in the table
            content = [[cell.value for cell in ent] for ent in data]

            header = content[0]
            rest = content[1:]

            df = pd.DataFrame(rest, columns=header)
            mapping[entry] = df
        
    return mapping


def example():
    example_path = Path('example data')
    input_filename_old = example_path / 'example-data-old.xlsx'
    input_filename_new = example_path / 'example-data-new.xlsx'
    all_dfs_old = get_all_tables_in_excel(input_filename_old)
    all_dfs_new = get_all_tables_in_excel(input_filename_new)

    # The following table names need to be defined in the Excel file
    df_old = all_dfs_old['Table_data']
    df_new = all_dfs_new['Table_data']
    output_filename = 'output_differences.xlsx'
    key_column = 'Part number'
    df_old_title = '5004-PL-2021 Rev 1 (example data)'
    df_new_title = '5004-PL-2021 Rev 2 (example data)'

    df_new.replace({np.nan: None}, inplace=True)
    df_old.replace({np.nan: None}, inplace=True)

    sp = SimilarPanda(new = df_new,
                      old = df_old,
                      key_column = key_column,
                      df_new_title = df_new_title,
                      df_old_title = df_old_title)
    sp.output_excel(output_filename)
    os.startfile(output_filename)
    print("done")

if __name__ == '__main__':
    print('Running the file directly gives sample data)')
    example()

